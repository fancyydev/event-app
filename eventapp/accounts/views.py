from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from .serializers import CustomUserSerializer, CustomRegisterSerializer
from rest_framework import status
from .models import CustomUser

from rest_framework.authtoken.models import Token
from django.shortcuts import get_object_or_404

from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication

from django.core.mail import send_mail

#Imports necesarios para recuperar contraseña mediante api
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from django.conf import settings
from django.urls import reverse

#Imports necesarios para crear el reporte
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch 
from events.models import Event
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import io

class Login(APIView):
    def post(self, request, format=None):
        user = get_object_or_404(CustomUser, email = request.data['email'])
        
        if not user.check_password(request.data['password']):
            return Response({"error" : "Invalid Password"}, status=status.HTTP_400_BAD_REQUEST)
        
        token, created = Token.objects.get_or_create(user = user)
        serializer = CustomUserSerializer(instance=user)
        
        return Response({"token" : token.key, "user" : serializer.data}, status=status.HTTP_200_OK)

    
class Register(APIView):
    def post(self, request, format = None):
        serializer = CustomRegisterSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            
            user = CustomUser.objects.get(email = serializer.data['email'])
            user.set_password(serializer.data['password'])
            user.save()
            
            token = Token.objects.create(user=user)
            
            #return Response({'token': token.key, "user" : serializer.data}, status=status.HTTP_201_CREATED)
            return Response(token.key, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class PasswordRecovery(APIView):
    def post(self, request, *args, **kwargs):
        email = request.data.get('email')

        # Verifica si el correo existe en la base de datos
        try:
            user = CustomUser.objects.get(email=email)
        except CustomUser.DoesNotExist:
            return Response({"detail": "User with this email does not exist."}, status=status.HTTP_400_BAD_REQUEST)

        # Genera el token de recuperación de contraseña
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))

        # Construye la URL de recuperación
        password_reset_url = request.build_absolute_uri(
            reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
        )

        # Renderiza el mensaje del correo
        context = {
            'user': user,
            'password_reset_url': password_reset_url,
        }
        email_subject = 'Password Reset Request'
        email_body = render_to_string('password_reset_email.html', context)

        # Envía el correo electrónico
        send_mail(
            email_subject,
            email_body,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            html_message=email_body 
        )

        return Response({"detail": "Password reset email has been sent."}, status=status.HTTP_200_OK)

class GenerateReportView(APIView):
    def get(self, request, event_id):
        # Obtener el evento
        try:
            event = Event.objects.get(id=event_id)
        except Event.DoesNotExist:
            return HttpResponse("Evento no encontrado", status=404)

        # Obtener usuarios registrados dentro de las fechas del evento
        users = CustomUser.objects.filter(
            created__date__range=[event.initial_date, event.end_date]
        ).values('email', 'name', 'phone', 'occupation', 'company',
                 'municipality__name', 'state__name', 'country__name', 'ticket')
        
        user_count = users.count()
        
        # Crear el PDF en memoria
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Agregar el título
        styles = getSampleStyleSheet()
        title_style = styles.get('Title', None)  # Obtener el estilo 'Title'
        if not title_style:
            return HttpResponse("Error al generar el estilo del título", status=500)
        
        title = Paragraph(f"Usuarios registrados durante el evento: {event.name_event}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.25 * inch))  # Espacio entre el título y la tabla

        body_style = styles.get('BodyText', None)
        if not body_style:
            return HttpResponse("Error al generar el estilo del cuerpo de la tabla", status=500)
        user_count_paragraph = Paragraph(
            f"Cantidad de usuarios registrados: {user_count}<br/><br/>Fecha: {event.initial_date} >< {event.end_date}",
            body_style
        )        
        elements.append(user_count_paragraph)
        elements.append(Spacer(1, 0.25 * inch))
        
        # Definir el ancho disponible para la tabla (ancho total de la página menos márgenes)
        page_width = letter[0] - 2 * inch  # ancho de la página menos márgenes de 1 pulgada en cada lado
        num_columns = 8
        column_width = page_width / num_columns

        # Crear los estilos de párrafo para las celdas
        cell_style = styles.get("BodyText", None)  # Obtener el estilo 'BodyText'
        if not cell_style:
            return HttpResponse("Error al generar el estilo del cuerpo de la tabla", status=500)
        
        cell_style.fontSize = 8  # Reducir el tamaño de la fuente
        cell_style.wordWrap = 'CJK'  # Permitir que el texto se divida en líneas

        # Crear la tabla de usuarios
        data = [
            ['Email', 'Name', 'Phone', 'Occupation', 'Company', 'Municipality', 'State', 'Country', 'Ticket']
        ]

        for user in users:
            data.append([
                Paragraph(user.get('email', '') or '', cell_style),
                Paragraph(user.get('name', '') or '', cell_style),
                Paragraph(user.get('phone', '') or '', cell_style),
                Paragraph(user.get('occupation', '') or '', cell_style),
                Paragraph(user.get('company', '') or '', cell_style),
                Paragraph(user.get('municipality__name', '') or '', cell_style),
                Paragraph(user.get('state__name', '') or '', cell_style),
                Paragraph(user.get('country__name', '') or '', cell_style),
                Paragraph(str(user.get('ticket', '')) or '', cell_style),
            ])

        table = Table(data, colWidths=[column_width] * num_columns)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), '#d0d0d0'),
            ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Asegura que el texto esté centrado verticalmente
        ]))

        elements.append(table)
        doc.build(elements)

        # Obtener el PDF en memoria y devolverlo como respuesta HTTP
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_usuarios_{event_id}.pdf"'
        return response
    
class GenerateReportExcelView(APIView):
    def get(self, request, event_id):
        # Obtener el evento
        try:
            event = Event.objects.get(id=event_id)
        except Event.DoesNotExist:
            return HttpResponse("Evento no encontrado", status=404)

        # Obtener usuarios registrados dentro de las fechas del evento
        users = CustomUser.objects.filter(
            created__date__range=[event.initial_date, event.end_date]
        ).values('email', 'name', 'phone', 'occupation', 'company',
                 'municipality__name', 'state__name', 'country__name', 'ticket')

        user_count = users.count()

        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Usuarios"

        # Establecer el título del evento en la primera celda
        title = f"Usuarios registrados durante el evento: {event.name_event}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(size=14, bold=True)  # Hacer el texto más grande y en negrita

        # Agregar la cantidad de usuarios registrados en la segunda celda
        count_info = f"Cantidad de usuarios registrados: {user_count}\nFecha: {event.initial_date} - {event.end_date}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        cell = ws.cell(row=2, column=1, value=count_info)
        cell.font = Font(size=12)  # Ajustar el tamaño de la fuente

        # Agregar un espacio en blanco para separar el título de la tabla
        ws.append([])

        # Escribir los encabezados de la tabla en la cuarta fila
        headers = ['Email', 'Name', 'Phone', 'Occupation', 'Company', 'Municipality', 'State', 'Country', 'ticket']
        ws.append(headers)

        # Escribir los datos de los usuarios
        for user in users:
            ws.append([
                user.get('email', ''),
                user.get('name', ''),
                user.get('phone', ''),
                user.get('occupation', ''),
                user.get('company', ''),
                user.get('municipality__name', ''),
                user.get('state__name', ''),
                user.get('country__name', ''),
                user.get('ticket', ''),
            ])

        # Ajustar el ancho de las columnas
        for i, column in enumerate(headers, 1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = 20  # Ajustar el ancho de las columnas según sea necesario

        # Preparar la respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_usuarios_{event_id}.xlsx"'

        # Guardar el libro de Excel en la respuesta
        wb.save(response)

        return response