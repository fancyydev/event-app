from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import CustomUser
import io

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch 


def generate_excel_report(event):
    # Obtener usuarios registrados dentro de las fechas del evento
    users = CustomUser.objects.filter(
        created__date__range=[event.initial_date, event.end_date]
    ).values('email', 'name', 'phone', 'occupation', 'company',
             'municipality__name', 'state__name', 'country__name', 'ticket')

    user_count = users.count()

    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Usuarios"

    # Establecer el título del evento en la primera celda
    title = f"Usuarios registrados durante el evento: {event.name_event}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(size=14, bold=True)  # Hacer el texto más grande y en negrita

    # Agregar la cantidad de usuarios registrados en la segunda celda
    count_info = f"Cantidad de usuarios registrados: {user_count}\nFecha: {event.initial_date} - {event.end_date}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    cell = ws.cell(row=2, column=1, value=count_info)
    cell.font = Font(size=12)  # Ajustar el tamaño de la fuente

    # Agregar un espacio en blanco para separar el título de la tabla
    ws.append([])

    # Escribir los encabezados de la tabla en la cuarta fila
    headers = ['Email', 'Name', 'Phone', 'Occupation', 'Company', 'Municipality', 'State', 'Country', 'ticket']
    ws.append(headers)

    # Escribir los datos de los usuarios
    for user in users:
        ws.append([
            user.get('email', ''),
            user.get('name', ''),
            user.get('phone', ''),
            user.get('occupation', ''),
            user.get('company', ''),
            user.get('municipality__name', ''),
            user.get('state__name', ''),
            user.get('country__name', ''),
            user.get('ticket', ''),
        ])

    # Ajustar el ancho de las columnas
    for i, column in enumerate(headers, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = 20  # Ajustar el ancho de las columnas según sea necesario

    # Guardar el archivo en un buffer en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_pdf_report(event):
    # Obtener usuarios registrados dentro de las fechas del evento
    users = CustomUser.objects.filter(
        created__date__range=[event.initial_date, event.end_date]
    ).values('email', 'name', 'phone', 'occupation', 'company',
             'municipality__name', 'state__name', 'country__name', 'ticket')
    
    user_count = users.count()
    
    # Crear el PDF en memoria
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Agregar el título
    styles = getSampleStyleSheet()
    title_style = styles.get('Title', None)
    if not title_style:
        raise ValueError("Error al generar el estilo del título")
    
    title = Paragraph(f"Usuarios registrados durante el evento: {event.name_event}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.25 * inch))  # Espacio entre el título y la tabla

    body_style = styles.get('BodyText', None)
    if not body_style:
        raise ValueError("Error al generar el estilo del cuerpo de la tabla")
    
    user_count_paragraph = Paragraph(
        f"Cantidad de usuarios registrados: {user_count}<br/><br/>Fecha: {event.initial_date} - {event.end_date}",
        body_style
    )
    elements.append(user_count_paragraph)
    elements.append(Spacer(1, 0.25 * inch))
    
    # Definir el ancho disponible para la tabla
    page_width = letter[0] - 2 * inch
    num_columns = 9
    column_width = page_width / num_columns

    # Crear los estilos de párrafo para las celdas
    cell_style = styles.get("BodyText", None)
    if not cell_style:
        raise ValueError("Error al generar el estilo del cuerpo de la tabla")
    
    cell_style.fontSize = 8
    cell_style.wordWrap = 'CJK'

    # Crear la tabla de usuarios
    data = [
        ['Email', 'Name', 'Phone', 'Occupation', 'Company', 'Municipality', 'State', 'Country', 'Ticket']
    ]

    for user in users:
        data.append([
            Paragraph(user.get('email', '') or '', cell_style),
            Paragraph(user.get('name', '') or '', cell_style),
            Paragraph(user.get('phone', '') or '', cell_style),
            Paragraph(user.get('occupation', '') or '', cell_style),
            Paragraph(user.get('company', '') or '', cell_style),
            Paragraph(user.get('municipality__name', '') or '', cell_style),
            Paragraph(user.get('state__name', '') or '', cell_style),
            Paragraph(user.get('country__name', '') or '', cell_style),
            Paragraph(str(user.get('ticket', '')) or '', cell_style),
        ])

    table = Table(data, colWidths=[column_width] * num_columns)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d0d0d0'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)
    doc.build(elements)

    # Obtener el PDF en memoria y devolverlo
    pdf = buffer.getvalue()
    buffer.close()

    return pdf
